from openpyxl import load_workbook

wb = load_workbook(r'D:\testfile\导入.xlsx')  # 生成一个已存在的wookbook对象
wb1 = wb.active  # 激活sheet


# 卡号
card_no = 89812604000
# 对应iccid
ICCID = 444489812604000

# 0,1 (0,第一次导入，1第二次导出)
is_import = 1
# 位置第几行
location = 1
if card_no < 89812605000:
    for i in range(0, 1000):
        card_no = card_no + 1
        ICCID = ICCID + 1
        for j in range(0, 1):
            location = location + 1
        wb1.cell(location, 1, card_no)
        wb1.cell(location, 2, ICCID)
        wb1.cell(location, 3, is_import)

# wb1.cell(2, 2, 'pass2')  # 往sheet中的第二行第二列写入‘pass2’的数据

wb.save(r'D:\testfile\导入2.xlsx')  # 保存
wb.close()
