import json
import pytest
import requests
from openpyxl import load_workbook
from tool import get_diff_login as gdl


wb = load_workbook('D:/Desktop/vv.xlsx', data_only=True)
sheet = wb.active

# 同一个卡，不同人的价格这一点计算正确
# 以下着测试下多个 不同套餐的情况下 计算结果是否正确
# 多卡-同套餐的卡 金额不对的问题 在公众号的充值流量这里出现过问题 所以现在续费处也许进行排查
# 单卡-月卡
Headers = {
    'content-type': 'application/json',
    'token': gdl.token_test13
}


def test_balance_month_card():
    data = json.dumps(data), headers=Headers)
    # # print(re.text)
    # allRenewPrice = re.json()['data']['allRenewPrice']
    # print(allRenewPrice)


# 多卡-月卡-同套餐
def test_balance_batch_month_card_same_pack():
    data = sheet.cell(row=2, column=4)
    re = requests.post(url=gdl.url, data=json.dumps(data), headers=Headers)
    allRenewPrice = re.json()['data']['allRenewPrice']
    print(allRenewPrice)


# 多卡-年卡-同套餐  1200m/月
def test_balance_batch_year_card_same_pack():
    data = sheet.cell(row=2, column=5)
    re = requests.post(url=gdl.url, data=json.dumps(data), headers=Headers)
    # print(re.text)
    allRenewPrice = re.json()['data']['allRenewPrice']
    print(allRenewPrice)
    ret = re.json()['ret']
    print(ret)
    assert allRenewPrice == 297.0 and ret == 1
    print('价格获取成功')
    sheet.cell(row=2, column=2)
    #     data_json = data.value
    #     print(data_json)
    #     re = requests.post(url=gdl.url, data=json.dumps(data_json), headers=Headers)
    #     print(re.text)
    #     print('l')
    #     allRenewPrice = re.json()['data']['allRenewPrice']
    #     print(123)
    #     print(allRenewPrice)
    #
    #
    # # 单卡-年卡
    # def test_balance_year_card():
    #     cell = sheet.cell(row=2, column=3)
    #     data = cell.value
    #     print(data)
    #     # re = requests.post(url=gdl.url, data

# 多卡-月卡-不同套餐 30M/月  5M/月
def test_balance_batch_month_card_diff_pack():
    data = sheet.cell(row=2, column=6)
    re = requests.post(url=gdl.url, data=json.dumps(data), headers=Headers)
    # print(re.text)
    allRenewPrice = re.json()['data']['allRenewPrice']
    print(allRenewPrice)
    ret = re.json()['ret']
    print(ret)
    assert allRenewPrice == 7.01 and ret == 1
    print('价格获取成功')


# 多卡-年卡-不同套餐  1200m/年 测试年包
def test_balance_batch_year_card_diff_pack():
    data = sheet.cell(row=2, column=7)
    re = requests.post(url=gdl.url, data=json.dumps(data), headers=Headers)
    # print(re.text)
    allRenewPrice = re.json()['data']['allRenewPrice']
    print(allRenewPrice)
    ret = re.json()['ret']
    print(ret)
    assert allRenewPrice == 149.01 and ret == 1
    print('价格获取成功')


# 多卡-年卡-月卡-混合-充值月  1200m/年 测试年包  30M/月  5M/月
def test_balance_batch_year_month():
    data = sheet.cell(row=2, column=8)
    re = requests.post(url=gdl.url, data=json.dumps(data), headers=Headers)
    # print(re.text)
    allRenewPrice = re.json()['data']['allRenewPrice']
    print(allRenewPrice)
    ret = re.json()['ret']
    print(ret)
    assert allRenewPrice == 233.11 and ret == 1
    print('价格获取成功')


# 多卡-年卡-月卡-混合-充年
def test_balance_batch_year_month_month():
    data = sheet.cell(row=2, column=9)
    re = requests.post(url=gdl.url, data=json.dumps(data), headers=Headers)
    # print(re.text)
    msg = re.json()['msg']
    assert msg == '年卡不能按月续费'
    print('测试通过')


if __name__ == '__main__':
    pytest.main("-vs", "get_package_price.py")

